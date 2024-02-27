import json
import time
import openpyxl
import requests
import math
from configparser import ConfigParser
import xlwings as xw
from time import strftime, localtime
import os
import pymssql

CONFIGFILE = '.\\config\\config_jstz.ini'
config = ConfigParser()
config.read(CONFIGFILE)

tuo = config['base']['tuo']
send_host = config['base']['send_host']
name = config['base']['name']
url = config['base']['get_url']


def get_sql(gysbm):
    try:
        sql = f'select sagegysbm from dbo.cgnxjhdmx where gysbm={gysbm}'
        wlmjson = requests.post('http://10.10.250.30:5002/getjsonall', data={"sql": sql}).text
        tmptask = json.loads(wlmjson)
        return tmptask[-1]['sagegysbm']
    except Exception:
        return '', ''


def get_sql1():
    try:
        db = pymssql.connect(host='10.10.100.165', user='HZGEARBOX', password='tiger', database='hcerp', autocommit=True, charset='UTF-8')
        curses = db.cursor()
        STRDAT_0 = input('请输入开始时间(格式:2024-01-01)：')
        ENDDAT_0 = input('请输入结束时间(格式:2024-01-01)：')
        sql = f"SELECT NUM_0 ,POHFCY_0 ,BPSNUM_0 ,STRDAT_0,ENDDAT_0 ,CREDAT_0  FROM YPURSTLH WHERE STRDAT_0='{STRDAT_0}' AND ENDDAT_0='{ENDDAT_0}'"
        # sql = f"SELECT NUM_0 ,POHFCY_0 ,BPSNUM_0 ,STRDAT_0,ENDDAT_0 ,CREDAT_0  FROM YPURSTLH WHERE STRDAT_0='2024-01-01' AND ENDDAT_0='2024-01-15'"
        curses.execute(sql)
        result = curses.fetchall()
        return result
    except Exception:
        return []


def s_f(len_str, lens, ws, rows):
    row = math.ceil(len_str/lens)
    if ws.row_dimensions[rows].height >= 20*row:
        return
    else:
        ws.row_dimensions[rows].height = 20*row


def print_file_(filename):
    try:
        app = xw.App(visible=False, add_book=False)  # 启动Excel程序
        workbook = app.books.open(filename)  # 打开要打印的工作簿
        workbook.api.PrintOut(Copies=1, Collate=True)  # 打印工作簿
        workbook.close()
        app.quit()
        print('开始打印...')
        print('打印成功')
    except Exception as e:
        print(e)
        print('打印出错！')


def main():
    try:
        res = requests.get(url=f'{url}{name}').text
        data = json.loads(res)
        # print(data)
    except Exception as e:
        # print(e)
        time.sleep(3)
        return

    if data != None:
        wb = openpyxl.load_workbook('./jstz_l.xlsx')
        ws = wb.active
        sheetnum = 0

        s1 = []
        for x in range(len(data)):
            gys_num = data[x]['LIFNR']
            s1.append(gys_num)
        s = []
        for i in s1:
            if i not in s:
                s.append(i)
        sheetnum += len(s)

        LI_names = ''
        for i in s:
            LI_names += '-' + i.lstrip("0")

        data_num = 0
        for tmprow in s:
            for x in range(len(data)):
                if data[x]['LIFNR'] == tmprow:
                    data_num += 1
            if data_num > 18:
                if data_num % 18 == 0:
                    sheetnum += (data_num // 18) - 1
                else:
                    sheetnum += data_num // 18
            data_num = 0
        for y in range(sheetnum - 1):
            newwb = wb.copy_worksheet(ws)
            newwb.title = f'Sheet{y+2}'

        t = 1
        lst1 = []
        for subtmprow in s:
            for x in range(len(data)):
                if data[x]['LIFNR'] == subtmprow:
                    t += 1
            lst1.append(t-1)
            t = 1
        dict1 = dict(zip(s, lst1))

        bm = get_sql1()

        m = 1
        m1 = 1
        s_num = 0
        n = 1
        n1 = 1
        nx = 1
        lst = []
        lst1 = []
        y_num = 1
        for tmpsubrow in dict1.items():
            j = 0
            j1 = 0
            if (tmpsubrow[1] % 18) == 0:
                Y_num = tmpsubrow[1]//18
            else:
                Y_num = tmpsubrow[1]//18 + 1
            for k in range(len(data)):
                ws = wb[f'Sheet{m}']
                time_end = strftime('%Y-%m-%d %H:%M:%S', localtime())
                ws['I45'].value = time_end
                ws['B45'].value = data[k]['FULLNAME']
                ws['B2'].value = data[k]['WERKS'] + '/' + data[k]['WERKS_T']
                ws['I2'].value = f'第{y_num}页，共{Y_num}页'
                tmptask = get_sql(tmpsubrow[0].lstrip("0"))
                # print(tmptask)
                for i in bm:
                    if i[2] == tmptask:
                        # print(i[0])
                        ws['K3'].value = i[0]
                if data[k]['LIFNR'] == tmpsubrow[0]:
                    ws['B3'].value = tmpsubrow[0].lstrip("0") + '/' + data[k]['NAME_ORG1'] + '/' + tmptask
                    lst.append(data[k]['MATNR'].lstrip("0"))
                    if data[k]['BUDAT_MKPF'][0:4] == '9999':
                        if data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = ' '
                        elif data[k]['PART_NO'] != '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        elif data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] != '':
                            ws[f'B{j + 6}'].value = data[k]['DRAWING_NO']
                        else:
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        ws[f'E{j + 6}'].value = data[k]['MATNR'].lstrip("0")
                        ws[f'E{j + 7}'].value = '小计数量：'
                        ws[f'G{j + 7}'].value = data[k]['ZXJ_SL']
                        ws[f'I{j + 7}'].value = '小计金额：'
                        ws[f'K{j + 7}'].value = data[k]['ZXJ_JE']
                        j += 2
                        n1 += 1
                    else:
                        if len(lst) == 1:
                            ws[f'A{j + 6}'].value = nx
                            nx += 1
                        else:
                            if lst[-1] != lst[-2]:
                                ws[f'A{j + 6}'].value = nx
                                nx += 1
                        JT_len = len(data[k]['PART_NO'] + '/' + data[k]['DRAWING_NO'])
                        s_f(JT_len, 40, ws, j + 6)
                        if data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = ' '
                        elif data[k]['PART_NO'] != '' and data[k]['DRAWING_NO'] == '':
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        elif data[k]['PART_NO'] == '' and data[k]['DRAWING_NO'] != '':
                            ws[f'B{j + 6}'].value = data[k]['DRAWING_NO']
                        else:
                            ws[f'B{j + 6}'].value = data[k]['PART_NO']
                        ws[f'E{j + 6}'].value = data[k]['MATNR'].lstrip("0")
                        ws[f'H{j + 6}'].value = data[k]['STPRS']
                        s_f(len(data[k]['MAKTX']), 35, ws, j + 7)
                        ws[f'B{j + 7}'].value = data[k]['MAKTX']
                        ws[f'E{j + 7}'].value = data[k]['MBLNR']
                        ws[f'G{j + 7}'].value = data[k]['MENGE']
                        ws[f'H{j + 7}'].value = data[k]['DMBTR']
                        # ws[f'I{j + 7}'].value = data[k]['EBELN']
                        ws[f'I{j + 6}'].value = data[k]['CHARG']
                        ws[f'K{j + 6}'].value = data[k]['BUDAT_MKPF']
                        # ws[f'K{j + 7}'].value = tmptask[1]
                        if data[k]['SOBKZ'] == '':
                            ws[f'G{j + 6}'].value = '非零'
                        else:
                            ws[f'G{j + 6}'].value = '零'
                        j += 2
                        n += 1
                        n1 += 1
                    if j + 6 > 40:
                        m += 1
                        j -= 36
                        y_num += 1
            MENGE_num = 0
            DMBTRL_num = 0
            DMBTRFL_num = 0
            lst1.append(n1-1)
            n_num1 = lst1[-1]
            for i in range(len(data)):
                ws = wb[f'Sheet{m1}']
                ws['H2'].value = nx-1
                if data[i]['LIFNR'] == tmpsubrow[0]:
                    if data[i]['BUDAT_MKPF'][0:4] != '9999':
                        MENGE_num += float(data[i]['MENGE'])
                        if data[i]['SOBKZ'] == '':
                            DMBTRFL_num += float(data[i]['DMBTR'])
                        else:
                            DMBTRL_num += float(data[i]['DMBTR'])
                        j1 += 2
                    else:
                        j1 += 2
                    if j1 + 6 > 40:
                        m1 += 1
                        j1 -= 36
                    if n_num1 <= 12:
                        ws['A42'].value = '非零库计划金额：'
                        ws['C42'].value = DMBTRFL_num
                        ws['G42'].value = '零库计划金额：'
                        ws['I42'].value = DMBTRL_num
                        ws['A43'].value = '数量合计：'
                        ws['C43'].value = MENGE_num
                        ws['G43'].value = '金额合计：'
                        ws['I43'].value = DMBTRL_num + DMBTRFL_num
            if n_num1 <= 18:
                s_num += 1
            else:
                if (n_num1 % 18) == 0:
                    s_num += (n_num1 // 18)
                    ws = wb[f'Sheet{s_num}']
                    ws['A42'].value = '非零库计划金额：'
                    ws['C42'].value = DMBTRFL_num
                    ws['G42'].value = '零库计划金额：'
                    ws['I42'].value = DMBTRL_num
                    ws['A43'].value = '数量合计：'
                    ws['C43'].value = MENGE_num
                    ws['G43'].value = '金额合计：'
                    ws['I43'].value = DMBTRL_num + DMBTRFL_num
                else:
                    s_num += (n_num1 // 18) + 1
                    ws = wb[f'Sheet{s_num}']
                    ws['A42'].value = '非零库计划金额：'
                    ws['C42'].value = DMBTRFL_num
                    ws['G42'].value = '零库计划金额：'
                    ws['I42'].value = DMBTRL_num
                    ws['A43'].value = '数量合计：'
                    ws['C43'].value = MENGE_num
                    ws['G43'].value = '金额合计：'
                    ws['I43'].value = DMBTRL_num + DMBTRFL_num
            m += 1
            m1 += 1
            n = 1
            n1 = 1
            y_num = 1
            nx = 1
            if tmpsubrow[1] % 18 == 0:
                m -= 1
                m1 -= 1
        wb.save(f'./结算通知单/jstz结算通知单{LI_names}.xlsx')
        # os.startfile(f'.\\结算通知单\\jstz结算通知单{LI_names}.xlsx')
    print(f'结算通知单输出成功(jstz结算通知单{LI_names})')
    print_file_(f'.\\结算通知单\\jstz结算通知单{LI_names}.xlsx')


if __name__ == "__main__":
    while True:
        try:
            main()
            time.sleep(1)
        except Exception as E:
            print(E)
